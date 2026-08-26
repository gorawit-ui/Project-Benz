import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Expense Tracking"

BRAND_FILL = PatternFill(start_color="1F5C4C", end_color="1F5C4C", fill_type="solid")
BRAND_FONT = Font(name="Arial", size=10.5, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F5C4C")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="6B6357")
BODY_FONT = Font(name="Arial", size=10.5)
EXAMPLE_FONT = Font(name="Arial", size=10, italic=True, color="6B6357")
MUSTHAVE_MARK_FONT = Font(name="Arial", size=8.5, bold=True, color="FFE9A8")

thin = Side(style="thin", color="D8D2C4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

groups = [
    ("ระบบจัดการ", "E9E5DA", "4A443A", [
        ("รหัสรายการ", True),
        ("วันที่บันทึกเข้าระบบ", True),
        ("ผู้บันทึก", True),
        ("สถานะ", True),
    ]),
    ("ประเภทเงิน", "D9EDE7", "1F5C4C", [
        ("ประเภทเงิน (เงินสดย่อย/เงินทดรองจ่าย)", True),
    ]),
    ("เอกสาร", "DCE8F5", "2F5C8A", [
        ("ประเภทเอกสาร", True),
        ("เลขที่เอกสาร", True),
        ("เลขที่ PO", False),
        ("วันที่ในบิล (Date)", True),
    ]),
    ("คู่ค้า", "FBF1DE", "8A5A1F", [
        ("ชื่อซัพพลายเออร์ (ไทย)", True),
        ("ชื่อซัพพลายเออร์ (English)", False),
        ("รายละเอียดค่าใช้จ่าย (Description)", True),
    ]),
    ("บัญชี (ตาม Odoo)", "EAE3F5", "5B3F8A", [
        ("หมวดหมู่ (ตาม Odoo)", True),
        ("Cost Center", False),
        ("Acc name", False),
    ]),
    ("ยอดเงิน", "E3F3EA", "1F7A45", [
        ("จำนวนเงินก่อน VAT", True),
        ("VAT 7%", True),
        ("ยอดรวม (Grand Total)", True),
    ]),
    ("ไฟล์ / หลักฐาน", "FDEBDD", "B4501F", [
        ("ลิงก์ไฟล์ใบเสร็จ (Drive)", True),
        ("ลิงก์เอกสารรับเงิน (กรณีบิลไม่สมบูรณ์)", False),
        ("แจ้งเตือนรายการซ้ำ", False),
    ]),
    ("เชื่อมต่อ Odoo", "FBE7EA", "A3324A", [
        ("ID Odoo / ID Express", False),
    ]),
    ("ตรวจทาน", "E7ECEF", "3D5566", [
        ("ผู้ตรวจทาน", False),
        ("วันที่ตรวจทาน", False),
        ("หมายเหตุ", False),
    ]),
]

flat_headers = []
for gname, fill, textcolor, cols in groups:
    for h, must in cols:
        flat_headers.append((gname, fill, textcolor, h, must))

n_cols = len(flat_headers)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
ws.cell(row=1, column=1, value="TDFB Expense Tracking — เบิกทดรองจ่าย / เงินสดย่อย").font = TITLE_FONT

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
ws.cell(row=2, column=1,
        value="คอลัมน์ 1-4 (ระบบจัดการ) กรอกจากระบบอัตโนมัติ ห้ามแก้มือ — หัวข้อที่มีเครื่องหมาย ★ คือกลุ่ม \"ต้องมีตั้งแต่รอบแรก\" ส่วนที่เหลือเพิ่มได้ทีหลัง"
        ).font = SUBTITLE_FONT

group_row = 4
header_row = 5
example_row = 6

col = 1
for gname, fill, textcolor, cols in groups:
    span = len(cols)
    ws.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=col + span - 1)
    c = ws.cell(row=group_row, column=col, value=gname)
    c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    c.font = Font(name="Arial", size=10, bold=True, color=textcolor)
    c.alignment = center
    for cc in range(col, col + span):
        ws.cell(row=group_row, column=cc).border = BORDER
        ws.cell(row=group_row, column=cc).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    col += span

for i, (gname, fill, textcolor, h, must) in enumerate(flat_headers, start=1):
    label = f"★ {h}" if must else h
    c = ws.cell(row=header_row, column=i, value=label)
    c.fill = BRAND_FILL
    c.font = BRAND_FONT
    c.alignment = center
    c.border = BORDER

example_values = [
    "EX-2026-0001", "12/08/2569 09:14", "พิมพ์วดี ว.", "รอตรวจ",
    "เงินทดรองจ่าย",
    "ใบเสร็จรับเงิน", "R-88213", "PO-0231", "12/08/2569",
    "ปตท. พหลโยธิน", "PTT Phaholyothin", "ค่าน้ำมันรถ เดินทางพบลูกค้า",
    "ค่าเดินทาง", "TD050100", "ค่าใช้จ่ายเดินทาง",
    1168.22, 81.78, 1250.00,
    "https://drive.google.com/... (ตัวอย่าง)", "", "ไม่ซ้ำ",
    "", "", "", "ตัวอย่างแถวนี้ — ลบก่อนใช้งานจริง",
]
for i, val in enumerate(example_values, start=1):
    c = ws.cell(row=example_row, column=i, value=val)
    c.font = EXAMPLE_FONT
    c.alignment = left_wrap
    c.border = BORDER

blank_rows = 20
NOTE_FILL = PatternFill(start_color="FCFBF8", end_color="FCFBF8", fill_type="solid")
for r in range(blank_rows):
    rr = example_row + 1 + r
    for i in range(1, n_cols + 1):
        c = ws.cell(row=rr, column=i, value="")
        c.font = BODY_FONT
        c.alignment = left_wrap
        c.border = BORDER
        c.fill = NOTE_FILL
    ws.row_dimensions[rr].height = 20

for i, (gname, fill, textcolor, h, must) in enumerate(flat_headers, start=1):
    letter = get_column_letter(i)
    length = len(h)
    ws.column_dimensions[letter].width = max(14, min(30, length * 0.95 + 4))

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 16
ws.row_dimensions[group_row].height = 20
ws.row_dimensions[header_row].height = 34
ws.row_dimensions[example_row].height = 30

ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

ws2 = wb.create_sheet("คำอธิบาย")
ws2.merge_cells("A1:C1")
ws2["A1"] = "คำอธิบายกลุ่มคอลัมน์"
ws2["A1"].font = TITLE_FONT

hdr2 = ["กลุ่ม", "ต้องมีตั้งแต่รอบแรก?", "หมายเหตุ"]
for i, h in enumerate(hdr2, start=1):
    c = ws2.cell(row=3, column=i, value=h)
    c.fill = BRAND_FILL
    c.font = BRAND_FONT
    c.alignment = center
    c.border = BORDER

legend_notes = {
    "ระบบจัดการ": "ระบบกรอกให้อัตโนมัติทั้งหมด ใช้จับคู่แถวตอนอัปเดตข้อมูล ห้ามแก้มือ",
    "ประเภทเงิน": "แยกเงินสดย่อย/เงินทดรองจ่าย เพื่อคำนวณ threshold ฿20,000/เดือน",
    "เอกสาร": "เลขที่ PO เพิ่มได้ทีหลังถ้ายังไม่พร้อม",
    "คู่ค้า": "ชื่ออังกฤษใช้ตั้งชื่อไฟล์ใน Drive เพิ่มได้ทีหลัง",
    "บัญชี (ตาม Odoo)": "Cost Center และ Acc name ต้องได้ผังบัญชีจริงจาก Odoo มาก่อนตั้งค่า",
    "ยอดเงิน": "แยกก่อน/หลัง VAT ให้คำนวณอัตโนมัติจาก Grand Total",
    "ไฟล์ / หลักฐาน": "ลิงก์เอกสารรับเงินใช้เฉพาะกรณีบิลไม่สมบูรณ์ (เช่น 7-11)",
    "เชื่อมต่อ Odoo": "กรอกมือหลังลงบัญชีใน Odoo แล้ว จนกว่าจะมีการเชื่อม API จริง",
    "ตรวจทาน": "เก็บไว้เป็น audit trail ว่าใครตรวจ เมื่อไหร่",
}

r = 4
for gname, fill, textcolor, cols in groups:
    must_any = any(m for _, m in cols)
    ws2.cell(row=r, column=1, value=gname).font = Font(name="Arial", size=10.5, bold=True)
    ws2.cell(row=r, column=1).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    ws2.cell(row=r, column=2, value="ต้องมี" if must_any else "เพิ่มได้ทีหลัง").font = BODY_FONT
    ws2.cell(row=r, column=3, value=legend_notes.get(gname, "")).font = BODY_FONT
    for cc in range(1, 4):
        ws2.cell(row=r, column=cc).border = BORDER
        ws2.cell(row=r, column=cc).alignment = left_wrap
    ws2.row_dimensions[r].height = 30
    r += 1

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 55

out_path = "TDFB Expense Tracking - Sheet Template.xlsx"
wb.save(out_path)
print("saved", out_path)
print("n_cols", n_cols)
