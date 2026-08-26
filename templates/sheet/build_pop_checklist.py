import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F5C4C", end_color="1F5C4C", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill(start_color="E7F0EC", end_color="E7F0EC", fill_type="solid")
GROUP_FONT = Font(name="Arial", size=10.5, bold=True, color="1F5C4C")
BODY_FONT = Font(name="Arial", size=10.5)
PRIORITY_HIGH_FONT = Font(name="Arial", size=10.5, bold=True, color="B7791F")
PRIORITY_MED_FONT = Font(name="Arial", size=10.5, color="6B6357")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F5C4C")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="6B6357")
NOTE_FILL = PatternFill(start_color="FCFBF8", end_color="FCFBF8", fill_type="solid")

thin = Side(style="thin", color="D8D2C4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------- Sheet 1 ----------
ws1 = wb.active
ws1.title = "คำถามคุยกับพี่ป๊อป"

ws1.merge_cells("A1:D1")
ws1["A1"] = "เตรียมคุยกับพี่ป๊อป — Expense Tracking (เงินสดย่อย + เงินทดรองจ่าย)"
ws1["A1"].font = TITLE_FONT
ws1.merge_cells("A2:D2")
ws1["A2"] = "จดคำตอบในคอลัมน์ 'คำตอบ/บันทึก' ระหว่างคุย — ใช้เป็นข้อมูลตั้งต้นออกแบบระบบต่อ"
ws1["A2"].font = SUBTITLE_FONT

headers = ["หมวด", "คำถาม", "คำตอบ/บันทึก", "ระดับความสำคัญ"]
header_row = 4
for col, text in enumerate(headers, start=1):
    c = ws1.cell(row=header_row, column=col, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = center
    c.border = BORDER

rows = [
    ("เงินสดย่อย", "วงเงินตั้งต้นจริงๆ เท่าไหร่ ใครถือ/รับผิดชอบ", "สำคัญมาก"),
    ("เงินสดย่อย", "ตอนนี้เติมเงินคืนกองทุนแบบไหน มีรอบตายตัวหรือเติมเมื่อใกล้หมด", "สำคัญมาก"),
    ("เงินสดย่อย", "ถ้าใช้เกินวงเงินก่อนถึงรอบเติม ตอนนี้จัดการยังไง", "ปานกลาง"),
    ("เงินทดรองจ่าย", "รอบจ่ายคืนพนักงานที่แท้จริงคือเมื่อไหร่", "สำคัญมาก"),
    ("เงินทดรองจ่าย", "จ่ายคืนยังไง โอนเข้าบัญชี / รวมในเงินเดือน / เงินสด", "สำคัญมาก"),
    ("เงินทดรองจ่าย", "มีขั้นตอนตรวจเอกสารที่ทำอยู่ตอนนี้ ใช้อะไรเป็นเกณฑ์ผ่าน/ไม่ผ่าน", "ปานกลาง"),
    ("หมวดหมู่ตาม Odoo", "ขอรายการหมวดหมู่/ผังบัญชีจริงที่ใช้กรอกอยู่ตอนนี้ (ขอ export หรือ screenshot)", "สำคัญมาก"),
    ("หมวดหมู่ตาม Odoo", "หมวดหมู่ของเงินสดย่อยกับเงินทดรองจ่ายใช้ชุดเดียวกันหรือคนละชุด", "สำคัญมาก"),
    ("Pain point และขนาดงาน", "ตอนนี้ทำเรื่องนี้เดือนละกี่ใบเสร็จ กี่คนที่เกี่ยวข้อง", "ปานกลาง"),
    ("Pain point และขนาดงาน", "ขั้นตอนไหนที่รู้สึกว่าเสียเวลา/ผิดพลาดบ่อยที่สุด ในคำพูดของตัวเอง", "สำคัญมาก"),
    ("Pain point และขนาดงาน", "ตอนนี้มีไฟล์ Excel/Google Sheet ที่ใช้สรุปอยู่แล้วไหม ถ้ามีขอไฟล์จริง", "สำคัญมาก"),
]

r = header_row + 1
group_start = r
prev_group = None
for group, question, priority in rows:
    ws1.cell(row=r, column=1, value=group).font = GROUP_FONT
    ws1.cell(row=r, column=1).alignment = wrap_left
    ws1.cell(row=r, column=1).border = BORDER
    ws1.cell(row=r, column=1).fill = GROUP_FILL

    qc = ws1.cell(row=r, column=2, value=question)
    qc.font = BODY_FONT
    qc.alignment = wrap_left
    qc.border = BORDER

    nc = ws1.cell(row=r, column=3, value="")
    nc.font = BODY_FONT
    nc.alignment = wrap_left
    nc.border = BORDER
    nc.fill = NOTE_FILL

    pc = ws1.cell(row=r, column=4, value=priority)
    pc.font = PRIORITY_HIGH_FONT if priority == "สำคัญมาก" else PRIORITY_MED_FONT
    pc.alignment = center
    pc.border = BORDER

    r += 1

def merge_group_column(ws, col, first_row, last_row):
    data = [ws.cell(row=rr, column=col).value for rr in range(first_row, last_row + 1)]
    start = first_row
    for i in range(1, len(data) + 1):
        if i == len(data) or data[i] != data[i - 1]:
            end = first_row + i - 1
            if end > start:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            start = first_row + i
            if i < len(data):
                start = first_row + i

merge_group_column(ws1, 1, group_start, r - 1)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 62
ws1.column_dimensions["C"].width = 40
ws1.column_dimensions["D"].width = 16
ws1.row_dimensions[1].height = 22
ws1.row_dimensions[2].height = 16
ws1.row_dimensions[header_row].height = 20
for rr in range(group_start, r):
    ws1.row_dimensions[rr].height = 34

ws1.freeze_panes = "A5"

# ---------- Sheet 2 ----------
ws2 = wb.create_sheet("หมวดหมู่ Odoo (กรอกระหว่างคุย)")

ws2.merge_cells("A1:D1")
ws2["A1"] = "รายการหมวดหมู่ Odoo — คัดลอกจากพี่ป๊อประหว่างคุย"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A2:D2")
ws2["A2"] = "กรอกทีละแถวตามที่พี่ป๊อปบอก แถวตัวอย่างด้านล่างเป็นตัวอย่างรูปแบบเท่านั้น ลบออกก่อนใช้งานจริง"
ws2["A2"].font = SUBTITLE_FONT

headers2 = ["รหัสบัญชี", "ชื่อหมวดหมู่", "ใช้กับ (เงินสดย่อย / เงินทดรองจ่าย / ทั้งคู่)", "หมายเหตุ"]
header_row2 = 4
for col, text in enumerate(headers2, start=1):
    c = ws2.cell(row=header_row2, column=col, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = center
    c.border = BORDER

example_row = header_row2 + 1
example_values = ["5210", "ค่าเดินทาง", "ทั้งคู่", "ตัวอย่าง — ลบแถวนี้ก่อนใช้งานจริง"]
EXAMPLE_FONT = Font(name="Arial", size=10.5, italic=True, color="6B6357")
for col, val in enumerate(example_values, start=1):
    c = ws2.cell(row=example_row, column=col, value=val)
    c.font = EXAMPLE_FONT
    c.alignment = wrap_left
    c.border = BORDER

blank_rows = 15
for i in range(blank_rows):
    rr = example_row + 1 + i
    for col in range(1, 5):
        c = ws2.cell(row=rr, column=col, value="")
        c.font = BODY_FONT
        c.alignment = wrap_left
        c.border = BORDER
        c.fill = NOTE_FILL

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 32
ws2.column_dimensions["D"].width = 34
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 16
ws2.row_dimensions[header_row2].height = 30

ws2.freeze_panes = "A5"

out_path = "คุยกับพี่ป๊อป_Expense Tracking.xlsx"
wb.save(out_path)
print("saved", out_path)
